"""XLSX workbook export.

Per spec, treat XLSX as a primary operational format. We use openpyxl to
produce stable, sortable, filterable workbooks. Column names are intentionally
boring; do not rename them casually.
"""

from __future__ import annotations

import json
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..core.models import State
from ..core.vault import Vault
from ..store import Store


@dataclass(frozen=True)
class _Profile:
    name: str
    description: str
    sheets: tuple[str, ...]
    include_states: tuple[str, ...]


PROFILES: dict[str, _Profile] = {
    "conversations_only": _Profile(
        name="conversations_only",
        description="One row per conversation. Useful for triage.",
        sheets=("conversations",),
        include_states=(State.PENDING_REVIEW.value, State.INDEXED.value, State.PRIVATE.value, State.QUARANTINED.value),
    ),
    "review_workbook": _Profile(
        name="review_workbook",
        description="Conversations + chunks for an analyst-friendly review pass.",
        sheets=("conversations", "chunks"),
        include_states=(State.PENDING_REVIEW.value, State.INDEXED.value, State.PRIVATE.value),
    ),
    "project_workbook": _Profile(
        name="project_workbook",
        description="Indexed-only conversations + chunks + projects + open loops + decisions.",
        sheets=("conversations", "chunks", "projects", "decisions", "open_loops", "entities", "groups", "provenance"),
        include_states=(State.INDEXED.value,),
    ),
    "full_analysis": _Profile(
        name="full_analysis",
        description="Everything that exists, except quarantined content (which has no derivatives).",
        sheets=("conversations", "chunks", "projects", "decisions", "open_loops", "entities", "preferences", "artifacts", "groups", "provenance"),
        include_states=(State.PENDING_REVIEW.value, State.INDEXED.value, State.PRIVATE.value, State.QUARANTINED.value),
    ),
}


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="404040")


def list_profiles() -> list[str]:
    return list(PROFILES.keys())


def export_workbook(
    vault: Vault, store: Store, *, profile: str = "review_workbook", out_path: Path | None = None
) -> Path:
    if profile not in PROFILES:
        raise ValueError(f"Unknown profile {profile!r}. Choices: {list(PROFILES)}")
    prof = PROFILES[profile]
    if out_path is None:
        ts = datetime.now().strftime("%Y%m%dT%H%M%S")
        out_path = vault.exports / f"{profile}_{ts}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # openpyxl creates a default sheet; we'll repurpose it for the first.
    default_ws = wb.active
    first_used = False

    for sheet_name in prof.sheets:
        builder = _SHEET_BUILDERS[sheet_name]
        if not first_used:
            ws = default_ws
            ws.title = sheet_name
            first_used = True
        else:
            ws = wb.create_sheet(title=sheet_name)
        builder(ws, store, prof.include_states)
        _finalize_sheet(ws)

    wb.save(out_path)
    return out_path


# --- formatting helpers -----------------------------------------------------

def _write_header(ws, columns: list[str]) -> None:
    for i, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"


def _finalize_sheet(ws) -> None:
    if ws.max_row >= 1 and ws.max_column >= 1:
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{max(ws.max_row, 1)}"
    # Crude column width: bound max width to keep things sortable.
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def _iso(ts: float | None) -> str:
    if ts is None:
        return ""
    try:
        return datetime.fromtimestamp(float(ts), tz=timezone.utc).isoformat()
    except (OverflowError, OSError, ValueError):
        return ""


def _state_filter_clause(states: tuple[str, ...]) -> tuple[str, list]:
    if not states:
        return "1=0", []
    placeholders = ",".join("?" for _ in states)
    return f"state IN ({placeholders})", list(states)


# --- sheet builders ---------------------------------------------------------

def _build_conversations(ws, store: Store, states: tuple[str, ...]) -> None:
    cols = [
        "conversation_id", "source", "title",
        "created_at", "updated_at", "imported_at",
        "state", "message_count",
        "summary_short", "summary_source", "manual_tags", "auto_tags",
        "primary_project_id", "importance_score", "resurfacing_score",
        "has_open_loops",
        "broad_group_label", "fine_group_label",
        "notes_local",
    ]
    _write_header(ws, cols)
    state_clause, params = _state_filter_clause(states)
    rows = store.conn.execute(
        f"""
        SELECT c.*,
               (SELECT COALESCE(g.llm_label, g.keyword_label)
                  FROM conversation_group_memberships m
                  JOIN conversation_groups g ON g.group_id = m.group_id
                 WHERE m.conversation_id = c.conversation_id AND g.level = 'broad'
                 LIMIT 1) AS broad_group_label,
               (SELECT COALESCE(g.llm_label, g.keyword_label)
                  FROM conversation_group_memberships m
                  JOIN conversation_groups g ON g.group_id = m.group_id
                 WHERE m.conversation_id = c.conversation_id AND g.level = 'fine'
                 LIMIT 1) AS fine_group_label
          FROM conversations c
         WHERE c.{state_clause}
         ORDER BY COALESCE(c.updated_at, c.created_at, c.imported_at) DESC
        """,
        params,
    ).fetchall()
    for r in rows:
        ws.append([
            r["conversation_id"], r["source"], r["title"],
            _iso(r["created_at"]), _iso(r["updated_at"]), _iso(r["imported_at"]),
            r["state"], r["message_count"],
            r["summary_short"] or "", r["summary_source"] or "deterministic",
            ", ".join(json.loads(r["manual_tags"] or "[]")),
            ", ".join(json.loads(r["auto_tags"] or "[]")),
            r["primary_project_id"] or "",
            r["importance_score"], r["resurfacing_score"],
            "yes" if r["has_open_loops"] else "no",
            r["broad_group_label"] or "",
            r["fine_group_label"] or "",
            r["notes_local"] or "",
        ])


def _build_chunks(ws, store: Store, states: tuple[str, ...]) -> None:
    cols = [
        "chunk_id", "conversation_id", "conversation_title", "chunk_index",
        "chunk_title", "start_message_ordinal", "end_message_ordinal",
        "summary_short", "project_id", "importance_score", "has_open_loops",
    ]
    _write_header(ws, cols)
    state_clause, params = _state_filter_clause(states)
    rows = store.conn.execute(
        f"""
        SELECT ch.*, c.title AS conv_title
          FROM chunks ch
          JOIN conversations c ON c.conversation_id = ch.conversation_id
         WHERE c.{state_clause}
         ORDER BY c.conversation_id, ch.chunk_index
        """,
        params,
    ).fetchall()
    for r in rows:
        ws.append([
            r["chunk_id"], r["conversation_id"], r["conv_title"], r["chunk_index"],
            r["chunk_title"] or "", r["start_message_ordinal"], r["end_message_ordinal"],
            r["summary_short"] or "", r["project_id"] or "",
            r["importance_score"], "yes" if r["has_open_loops"] else "no",
        ])


def _build_kind(ws, store: Store, kind: str, states: tuple[str, ...]) -> None:
    cols = [
        "object_id", "kind", "title", "description", "project_id",
        "source_conversation_ids", "source_chunk_ids", "current_state",
    ]
    _write_header(ws, cols)
    state_clause, params = _state_filter_clause(states)
    rows = store.conn.execute(
        f"""
        SELECT o.*,
               GROUP_CONCAT(DISTINCT p.conversation_id) AS conv_ids,
               GROUP_CONCAT(DISTINCT IFNULL(p.chunk_id, '')) AS chunk_ids
          FROM derived_objects o
          JOIN provenance_links p ON p.object_id = o.object_id
          JOIN conversations c ON c.conversation_id = p.conversation_id
         WHERE o.kind = ? AND o.state = 'active' AND c.{state_clause}
         GROUP BY o.object_id
         ORDER BY o.title
        """,
        [kind, *params],
    ).fetchall()
    for r in rows:
        ws.append([
            r["object_id"], r["kind"], r["title"], r["description"] or "",
            r["project_id"] or "", r["conv_ids"] or "",
            (r["chunk_ids"] or "").strip(","), r["state"],
        ])


def _build_projects(ws, store, states):    _build_kind(ws, store, "project", states)
def _build_decisions(ws, store, states):   _build_kind(ws, store, "decision", states)
def _build_open_loops(ws, store, states):  _build_kind(ws, store, "open_loop", states)
def _build_entities(ws, store, states):    _build_kind(ws, store, "entity", states)
def _build_preferences(ws, store, states): _build_kind(ws, store, "preference", states)
def _build_artifacts(ws, store, states):   _build_kind(ws, store, "artifact", states)


def _build_provenance(ws, store: Store, states: tuple[str, ...]) -> None:
    cols = [
        "link_id", "object_id", "object_kind", "object_title",
        "conversation_id", "conversation_title", "chunk_id", "excerpt",
    ]
    _write_header(ws, cols)
    state_clause, params = _state_filter_clause(states)
    rows = store.conn.execute(
        f"""
        SELECT p.*, o.kind AS object_kind, o.title AS object_title,
               c.title AS conv_title
          FROM provenance_links p
          JOIN derived_objects o ON o.object_id = p.object_id
          JOIN conversations c ON c.conversation_id = p.conversation_id
         WHERE c.{state_clause}
         ORDER BY o.kind, o.title, c.conversation_id
        """,
        params,
    ).fetchall()
    for r in rows:
        ws.append([
            r["link_id"], r["object_id"], r["object_kind"], r["object_title"],
            r["conversation_id"], r["conv_title"], r["chunk_id"] or "",
            (r["excerpt"] or "")[:500],
        ])


def _build_groups(ws, store: Store, states: tuple[str, ...]) -> None:
    cols = [
        "group_id", "level", "keyword_label", "llm_label", "member_count",
        "member_ids",
    ]
    _write_header(ws, cols)
    rows = store.conn.execute(
        """
        SELECT g.group_id, g.level, g.keyword_label, g.llm_label, g.member_count,
               GROUP_CONCAT(m.conversation_id, ',') AS member_ids
          FROM conversation_groups g
          LEFT JOIN conversation_group_memberships m ON m.group_id = g.group_id
         GROUP BY g.group_id
         ORDER BY g.level, g.member_count DESC
        """
    ).fetchall()
    for r in rows:
        ws.append([
            r["group_id"], r["level"],
            r["keyword_label"] or "", r["llm_label"] or "",
            r["member_count"], r["member_ids"] or "",
        ])


_SHEET_BUILDERS = {
    "conversations": _build_conversations,
    "chunks": _build_chunks,
    "projects": _build_projects,
    "decisions": _build_decisions,
    "open_loops": _build_open_loops,
    "entities": _build_entities,
    "preferences": _build_preferences,
    "artifacts": _build_artifacts,
    "groups": _build_groups,
    "provenance": _build_provenance,
}
