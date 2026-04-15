"""XLSX export smoke tests.

We don't render and visually inspect; we just verify the workbook exists,
has the expected sheets, and that the sheets contain the expected rows.
"""

from __future__ import annotations

from openpyxl import load_workbook

from threadatlas.core.models import State
from threadatlas.core.workflow import transition_state
from threadatlas.export import PROFILES, export_workbook
from threadatlas.extract import chunk_conversation, extract_for_conversation
from threadatlas.ingest import import_path


def _seed(tmp_vault, store, factory):
    path = factory([
        {"title": "Project Atlas roadmap",
         "messages": [
             ("user", "Project Atlas. I will pick option B.", 1.0),
             ("assistant", "Agreed. We agreed to ship in April.", 2.0),
             ("user", "TODO: revisit staffing.", 3.0),
             ("assistant", "Noted.", 4.0),
         ]},
    ])
    res = import_path(tmp_vault, store, path)
    cid = res.imported[0]
    transition_state(store, cid, State.INDEXED.value)
    chunk_conversation(store, cid)
    extract_for_conversation(store, cid)
    return cid


def test_review_workbook_has_expected_sheets(tmp_vault, store, chatgpt_export_factory):
    _seed(tmp_vault, store, chatgpt_export_factory)
    out = export_workbook(tmp_vault, store, profile="review_workbook")
    assert out.exists()
    wb = load_workbook(out)
    assert set(wb.sheetnames) == set(PROFILES["review_workbook"].sheets)


def test_project_workbook_contains_provenance_rows(tmp_vault, store, chatgpt_export_factory):
    _seed(tmp_vault, store, chatgpt_export_factory)
    out = export_workbook(tmp_vault, store, profile="project_workbook")
    wb = load_workbook(out)
    prov = wb["provenance"]
    # header + at least one provenance row from extraction
    assert prov.max_row >= 2


def test_workbook_filters_by_profile_visibility(tmp_vault, store, chatgpt_export_factory):
    """project_workbook is indexed-only; pending_review conv must not appear."""
    factory = chatgpt_export_factory
    path = factory([
        {"title": "Visible project thread",
         "messages": [("user", "I will pick A", 1.0), ("assistant", "ok", 2.0),
                      ("user", "more", 3.0), ("assistant", "ok", 4.0)]},
        {"title": "Hidden pending thread",
         "messages": [("user", "secret", 1.0), ("assistant", "ok", 2.0)]},
    ])
    res = import_path(tmp_vault, store, path)
    cid_visible, _cid_hidden = res.imported
    transition_state(store, cid_visible, State.INDEXED.value)
    chunk_conversation(store, cid_visible)
    extract_for_conversation(store, cid_visible)
    out = export_workbook(tmp_vault, store, profile="project_workbook")
    wb = load_workbook(out)
    rows = list(wb["conversations"].iter_rows(values_only=True))
    # Header + exactly one (visible) row.
    assert len(rows) == 2
    assert rows[1][0] == cid_visible
