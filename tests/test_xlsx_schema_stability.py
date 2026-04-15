"""XLSX column-header stability tests.

Column names are part of the contract: analysts script against them.
If you need to change a column, update this test in the same PR so the
change is explicit.
"""

from __future__ import annotations

from openpyxl import load_workbook

from threadatlas.core.models import State
from threadatlas.core.workflow import transition_state
from threadatlas.export import export_workbook
from threadatlas.extract import chunk_conversation, extract_for_conversation
from threadatlas.ingest import import_path


EXPECTED_COLUMNS = {
    "conversations": [
        "conversation_id", "source", "title",
        "created_at", "updated_at", "imported_at",
        "state", "message_count",
        "summary_short", "summary_source", "manual_tags", "auto_tags",
        "primary_project_id", "importance_score", "resurfacing_score",
        "has_open_loops",
        "broad_group_label", "fine_group_label",
        "notes_local",
    ],
    "chunks": [
        "chunk_id", "conversation_id", "conversation_title", "chunk_index",
        "chunk_title", "start_message_ordinal", "end_message_ordinal",
        "summary_short", "project_id", "importance_score", "has_open_loops",
    ],
    "projects": [
        "object_id", "kind", "title", "description", "project_id",
        "source_conversation_ids", "source_chunk_ids", "current_state",
    ],
    "decisions": [
        "object_id", "kind", "title", "description", "project_id",
        "source_conversation_ids", "source_chunk_ids", "current_state",
    ],
    "open_loops": [
        "object_id", "kind", "title", "description", "project_id",
        "source_conversation_ids", "source_chunk_ids", "current_state",
    ],
    "entities": [
        "object_id", "kind", "title", "description", "project_id",
        "source_conversation_ids", "source_chunk_ids", "current_state",
    ],
    "preferences": [
        "object_id", "kind", "title", "description", "project_id",
        "source_conversation_ids", "source_chunk_ids", "current_state",
    ],
    "artifacts": [
        "object_id", "kind", "title", "description", "project_id",
        "source_conversation_ids", "source_chunk_ids", "current_state",
    ],
    "provenance": [
        "link_id", "object_id", "object_kind", "object_title",
        "conversation_id", "conversation_title", "chunk_id", "excerpt",
    ],
}


def _seed(tmp_vault, store, factory):
    path = factory([
        {"title": "Project Theta roadmap",
         "messages": [
             ("user", "Project Theta. I decided to go with option B.", 1.0),
             ("assistant", "We agreed to ship in April.", 2.0),
             ("user", "TODO: revisit staffing.", 3.0),
             ("assistant", "Noted.", 4.0),
         ]},
    ])
    res = import_path(tmp_vault, store, path)
    cid = res.imported[0]
    transition_state(store, cid, State.INDEXED.value)
    chunk_conversation(store, cid)
    extract_for_conversation(store, cid)
    return cid


def test_review_workbook_columns_exact(tmp_vault, store, chatgpt_export_factory):
    _seed(tmp_vault, store, chatgpt_export_factory)
    out = export_workbook(tmp_vault, store, profile="review_workbook")
    wb = load_workbook(out)
    for sheet in ("conversations", "chunks"):
        headers = [c.value for c in wb[sheet][1]]
        assert headers == EXPECTED_COLUMNS[sheet], f"{sheet} columns drifted: {headers}"


def test_project_workbook_columns_exact(tmp_vault, store, chatgpt_export_factory):
    _seed(tmp_vault, store, chatgpt_export_factory)
    out = export_workbook(tmp_vault, store, profile="project_workbook")
    wb = load_workbook(out)
    for sheet in ("conversations", "chunks", "projects", "decisions",
                  "open_loops", "entities", "provenance"):
        headers = [c.value for c in wb[sheet][1]]
        assert headers == EXPECTED_COLUMNS[sheet], f"{sheet} columns drifted: {headers}"


def test_full_analysis_columns_exact(tmp_vault, store, chatgpt_export_factory):
    _seed(tmp_vault, store, chatgpt_export_factory)
    out = export_workbook(tmp_vault, store, profile="full_analysis")
    wb = load_workbook(out)
    for sheet, expected in EXPECTED_COLUMNS.items():
        if sheet not in wb.sheetnames:
            continue
        headers = [c.value for c in wb[sheet][1]]
        assert headers == expected, f"{sheet} columns drifted: {headers}"
