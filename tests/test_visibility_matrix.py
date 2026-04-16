"""Visibility matrix: state x surface invariants.

Enumerates every visibility state and every surface where content could
leak, then asserts the expected observable behavior. If a new state or
surface is added, this test must be updated.

The matrix (content source state -> surfaces):

    pending_review: {cli-search: NO, cli-search --include-private: NO,
                     mcp: NO, project synthesis: NO,
                     export review_workbook: YES, export project_workbook: NO,
                     fts rows: NO}
    indexed:        {cli-search: YES, cli-search --include-private: YES,
                     mcp: YES, project synthesis: YES,
                     export review_workbook: YES, export project_workbook: YES,
                     fts rows: YES}
    private:        {cli-search: NO, cli-search --include-private: YES,
                     mcp: NO, project synthesis: NO,
                     export review_workbook: YES, export project_workbook: NO,
                     fts rows: YES}
    quarantined:    {cli-search: NO, cli-search --include-private: NO,
                     mcp: NO, project synthesis: NO,
                     export review_workbook: NO (not in review_workbook states),
                     export project_workbook: NO,
                     fts rows: NO, chunks: NO, provenance: NO}
"""

from __future__ import annotations

from openpyxl import load_workbook

from threadatlas.core.models import MCP_VISIBLE_STATES, State
from threadatlas.core.workflow import transition_state
from threadatlas.export import export_workbook
from threadatlas.extract import chunk_conversation, extract_for_conversation
from threadatlas.ingest import import_path
from threadatlas.mcp.server import build_tools
from threadatlas.search import search_conversations


UNIQUE_TOKEN = "matrixuniqueterm"


def _seed_all_states(tmp_vault, store, factory):
    """Create one conversation per state with the same unique search token."""
    convs = []
    for label in ["A", "B", "C", "D"]:
        convs.append({
            "title": f"Matrix thread {label}",
            "messages": [
                ("user", f"{UNIQUE_TOKEN} discussion in thread {label}. I prefer option {label}.", 1.0),
                ("assistant", f"Got it. {UNIQUE_TOKEN} noted.", 2.0),
                ("user", "TODO: revisit next week.", 3.0),
                ("assistant", "Noted.", 4.0),
            ],
        })
    path = factory(convs)
    res = import_path(tmp_vault, store, path)
    a, b, c, d = res.imported
    # a -> pending_review (default); b -> indexed; c -> private; d -> quarantined
    transition_state(store, b, State.INDEXED.value)
    chunk_conversation(store, b)
    extract_for_conversation(store, b)
    transition_state(store, c, State.PRIVATE.value)
    chunk_conversation(store, c)
    extract_for_conversation(store, c)
    transition_state(store, d, State.QUARANTINED.value)
    return {"pending_review": a, "indexed": b, "private": c, "quarantined": d}


def _fts_row_count(store, table: str, conversation_id: str) -> int:
    """Count FTS rows whose rowid belongs to this conversation's rows.

    Uses the fact that our FTS tables share rowids with their source tables.
    """
    if table == "fts_conversations":
        row = store.conn.execute(
            "SELECT COUNT(*) AS c FROM fts_conversations WHERE rowid IN (SELECT rowid FROM conversations WHERE conversation_id = ?)",
            (conversation_id,),
        ).fetchone()
    elif table == "fts_messages":
        row = store.conn.execute(
            "SELECT COUNT(*) AS c FROM fts_messages WHERE rowid IN (SELECT rowid FROM messages WHERE conversation_id = ?)",
            (conversation_id,),
        ).fetchone()
    elif table == "fts_chunks":
        row = store.conn.execute(
            "SELECT COUNT(*) AS c FROM fts_chunks WHERE rowid IN (SELECT rowid FROM chunks WHERE conversation_id = ?)",
            (conversation_id,),
        ).fetchone()
    else:
        raise KeyError(table)
    return row["c"]


# ---------------------------------------------------------------------------
# Search surfaces
# ---------------------------------------------------------------------------

def test_cli_default_search_excludes_non_indexed(tmp_vault, store, chatgpt_export_factory):
    ids = _seed_all_states(tmp_vault, store, chatgpt_export_factory)
    hits = search_conversations(store, UNIQUE_TOKEN, visible_states=tuple(MCP_VISIBLE_STATES))
    returned = {h.conversation_id for h in hits}
    assert ids["indexed"] in returned
    assert ids["pending_review"] not in returned
    assert ids["private"] not in returned
    assert ids["quarantined"] not in returned


def test_cli_include_private_still_excludes_pending_and_quarantined(tmp_vault, store, chatgpt_export_factory):
    ids = _seed_all_states(tmp_vault, store, chatgpt_export_factory)
    visible = (State.INDEXED.value, State.PRIVATE.value)
    hits = search_conversations(store, UNIQUE_TOKEN, visible_states=visible)
    returned = {h.conversation_id for h in hits}
    assert ids["indexed"] in returned
    assert ids["private"] in returned
    assert ids["pending_review"] not in returned
    assert ids["quarantined"] not in returned


# ---------------------------------------------------------------------------
# MCP surface
# ---------------------------------------------------------------------------

def test_mcp_returns_indexed_only(tmp_vault, store, chatgpt_export_factory):
    ids = _seed_all_states(tmp_vault, store, chatgpt_export_factory)
    tools = build_tools(tmp_vault, store)
    import json as _json
    result = tools["query"].fn({"query": UNIQUE_TOKEN, "limit": 20})
    payload = _json.loads(result["content"][0]["text"])
    conv_hits = [h for h in payload.get("hits", []) if h["hit_type"] == "conversation"]
    returned = {h["id"] for h in conv_hits}
    assert returned == {ids["indexed"]}


def test_mcp_refuses_non_indexed_summary(tmp_vault, store, chatgpt_export_factory):
    ids = _seed_all_states(tmp_vault, store, chatgpt_export_factory)
    tools = build_tools(tmp_vault, store)
    for key in ("pending_review", "private", "quarantined"):
        r = tools["get_conversation_summary"].fn({"conversation_id": ids[key]})
        assert r.get("isError") is True, f"{key} should be hidden from MCP"


# ---------------------------------------------------------------------------
# FTS surface (defense-in-depth: no rows for pending_review or quarantined)
# ---------------------------------------------------------------------------

def test_fts_rows_only_for_fts_indexed_states(tmp_vault, store, chatgpt_export_factory):
    ids = _seed_all_states(tmp_vault, store, chatgpt_export_factory)
    # indexed and private: have FTS rows
    for key in ("indexed", "private"):
        assert _fts_row_count(store, "fts_conversations", ids[key]) == 1, key
        assert _fts_row_count(store, "fts_messages", ids[key]) > 0, key
    # pending_review and quarantined: NO FTS rows at all
    for key in ("pending_review", "quarantined"):
        assert _fts_row_count(store, "fts_conversations", ids[key]) == 0, key
        assert _fts_row_count(store, "fts_messages", ids[key]) == 0, key
        assert _fts_row_count(store, "fts_chunks", ids[key]) == 0, key


def test_fts_rows_rebuilt_respect_state(tmp_vault, store, chatgpt_export_factory):
    ids = _seed_all_states(tmp_vault, store, chatgpt_export_factory)
    store.rebuild_all_fts()
    store.conn.commit()
    # After full rebuild, invariant still holds.
    for key in ("pending_review", "quarantined"):
        assert _fts_row_count(store, "fts_messages", ids[key]) == 0
    for key in ("indexed", "private"):
        assert _fts_row_count(store, "fts_messages", ids[key]) > 0


# ---------------------------------------------------------------------------
# Export surface
# ---------------------------------------------------------------------------

def test_project_workbook_indexed_only(tmp_vault, store, chatgpt_export_factory):
    ids = _seed_all_states(tmp_vault, store, chatgpt_export_factory)
    out = export_workbook(tmp_vault, store, profile="project_workbook")
    wb = load_workbook(out)
    rows = list(wb["conversations"].iter_rows(values_only=True))[1:]
    seen = {r[0] for r in rows}
    assert ids["indexed"] in seen
    for key in ("pending_review", "private", "quarantined"):
        assert ids[key] not in seen


def test_review_workbook_excludes_quarantined(tmp_vault, store, chatgpt_export_factory):
    ids = _seed_all_states(tmp_vault, store, chatgpt_export_factory)
    out = export_workbook(tmp_vault, store, profile="review_workbook")
    wb = load_workbook(out)
    rows = list(wb["conversations"].iter_rows(values_only=True))[1:]
    seen = {r[0] for r in rows}
    # review_workbook includes pending_review + indexed + private but not quarantined
    assert ids["pending_review"] in seen
    assert ids["indexed"] in seen
    assert ids["private"] in seen
    assert ids["quarantined"] not in seen


# ---------------------------------------------------------------------------
# Chunk/provenance stripping under quarantine
# ---------------------------------------------------------------------------

def test_quarantined_has_no_chunks_or_provenance(tmp_vault, store, chatgpt_export_factory):
    ids = _seed_all_states(tmp_vault, store, chatgpt_export_factory)
    q = ids["quarantined"]
    assert store.list_chunks(q) == []
    assert store.list_provenance_for_conversation(q) == []
