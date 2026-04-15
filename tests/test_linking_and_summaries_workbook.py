"""Project linking + summaries_review XLSX profile."""

from __future__ import annotations

from openpyxl import load_workbook

from threadatlas.core.models import State
from threadatlas.core.workflow import transition_state
from threadatlas.export import export_workbook, PROFILES
from threadatlas.extract import chunk_conversation, extract_for_conversation
from threadatlas.ingest import import_path


def _seed(tmp_vault, store, factory):
    path = factory([
        {"title": "Project Atlas kickoff", "messages": [
            ("user", "Project Atlas. I will pick option B.", 1.0),
            ("assistant", "ok", 2.0),
            ("user", "TODO: confirm staffing.", 3.0),
            ("assistant", "ok", 4.0),
        ]},
    ])
    res = import_path(tmp_vault, store, path)
    cid = res.imported[0]
    transition_state(store, cid, State.INDEXED.value, vault=tmp_vault)
    chunk_conversation(store, cid)
    extract_for_conversation(store, cid)
    return cid


def test_link_sets_primary_project_id(tmp_vault, store, chatgpt_export_factory):
    cid = _seed(tmp_vault, store, chatgpt_export_factory)
    proj = next(o for o in store.list_derived_objects(kind="project"))
    store.update_conversation_meta(cid, primary_project_id=proj.object_id)
    store.conn.commit()
    c = store.get_conversation(cid)
    assert c.primary_project_id == proj.object_id


def test_unlink_clears(tmp_vault, store, chatgpt_export_factory):
    cid = _seed(tmp_vault, store, chatgpt_export_factory)
    proj = next(o for o in store.list_derived_objects(kind="project"))
    store.update_conversation_meta(cid, primary_project_id=proj.object_id)
    store.conn.commit()
    store.conn.execute(
        "UPDATE conversations SET primary_project_id = NULL WHERE conversation_id = ?",
        (cid,),
    )
    store.conn.commit()
    assert store.get_conversation(cid).primary_project_id is None


def test_summaries_review_profile_is_registered():
    assert "summaries_review" in PROFILES


def test_summaries_review_workbook_columns_exact(tmp_vault, store, chatgpt_export_factory):
    _seed(tmp_vault, store, chatgpt_export_factory)
    out = export_workbook(tmp_vault, store, profile="summaries_review")
    wb = load_workbook(out)
    assert wb.sheetnames == ["summaries_review"]
    headers = [c.value for c in wb["summaries_review"][1]]
    expected = [
        "conversation_id", "state", "source", "title",
        "summary_source", "summary_short",
        "first_user_message",
        "message_count", "chunk_count",
        "operator_decision", "operator_note",
    ]
    assert headers == expected


def test_summaries_review_workbook_contains_first_user_message(tmp_vault, store, chatgpt_export_factory):
    _seed(tmp_vault, store, chatgpt_export_factory)
    out = export_workbook(tmp_vault, store, profile="summaries_review")
    wb = load_workbook(out)
    ws = wb["summaries_review"]
    # Header + at least one data row.
    assert ws.max_row >= 2
    # first_user_message column (index 7, 1-based) should be populated.
    first_user_idx = 7
    row2 = [ws.cell(row=2, column=i).value for i in range(1, ws.max_column + 1)]
    assert "Project Atlas" in (row2[first_user_idx - 1] or "")
    # operator_decision (col 10, 1-based) should be blank by default.
    assert row2[9] in (None, "")
